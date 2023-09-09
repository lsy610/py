# import openpyxl
from openpyxl import Workbook, load_workbook
wb=Workbook()
ws=wb.active
ws.title='qq'#更改工作表名稱
ws.append([123,456,789,0])
wb.save('write.xlsx')
# wb=load_workbook('excel.xlsx')#excel路徑
# ws=wb.active#開啟預設工作表(工作表1)
# ws=wb['輸入要設定的工作表']
# wb.create_sheet('qq')新增工作表
# print(wb.sheetnames)查看總共有幾個工作表

# ws['A5'].value="小灰" #將A5儲存格改成小灰
# wb.save('excel.xlsx') 儲存檔案
# print(ws['A5'].value)